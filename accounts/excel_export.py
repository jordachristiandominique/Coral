"""Excel (.xlsx) export for a single survey batch.

Reproduces the CPCE COREMAP "Data summary" layout: benthic categories run down
the rows, each analyzed image is a column ("transect"), and every cell is that
category's percentage of the transect. MEAN, STD. DEV. (sample, n-1) and STD.
ERROR columns summarize across the images.

Uses the project's seven benthic classes as the Major Categories. The
"Total points (minus tape+wand+shadow)" row and the Subcategories section from
the original CPCE sheet are intentionally out of scope for now.
"""
from io import BytesIO
from statistics import mean as _mean, stdev as _stdev
import math
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# The seven benthic "major categories" in fixed order, shown in the CPCE style
# (UPPER CASE with the code in parentheses) to match the sample sheet.
CLASS_ORDER = [
    'Hard Coral', 'Soft Coral', 'Macroalgae', 'Halimeda',
    'Algae Assemblage', 'Abiotic', 'Other Biota',
]
CATEGORY_DISPLAY = {
    'Hard Coral': 'HARD CORAL (HC)',
    'Soft Coral': 'SOFT CORAL (SC)',
    'Macroalgae': 'MACROALGAE (MA)',
    'Halimeda': 'HALIMEDA (HA)',
    'Algae Assemblage': 'ALGAE ASSEMBLAGE (AA)',
    'Abiotic': 'ABIOTIC (AB)',
    'Other Biota': 'OTHER BIOTA (OB)',
}

# ---- Styling ----
TEAL = '1C5F6D'
GRAY = 'D9D9D9'          # zebra fill for alternate category rows
HEADER_GRAY = 'BFBFBF'   # section-header fill
NUM_FMT = '0.00'

_LABEL = Font(name='Calibri', size=10, bold=True, color='000000')
_PLAIN = Font(name='Calibri', size=10, color='000000')
_HEADER = Font(name='Calibri', size=10, bold=True, color='000000')
_LEFT = Alignment(horizontal='left', vertical='center')
_CENTER = Alignment(horizontal='center', vertical='center')

_thin = Side(style='thin', color='808080')
_med = Side(style='medium', color='000000')
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _safe_filename(batch):
    base = re.sub(r'[^A-Za-z0-9]+', '_', (batch.name or 'batch')).strip('_') or 'batch'
    return f'{base}_coral_coverage.xlsx'


def _pct_series(images):
    """Return {class_name: [pct_per_image, ...]} and per-image totals."""
    series = {name: [] for name in CLASS_ORDER}
    totals = []
    for image in images:
        pcs = image.point_classes or []
        total = len(pcs)
        totals.append(total)
        for name in CLASS_ORDER:
            count = sum(1 for p in pcs if p == name)
            series[name].append((count / total * 100) if total else 0.0)
    return series, totals


def _stats(values):
    """MEAN, sample STD. DEV. (n-1), and STD. ERROR for a category's per-image %."""
    n = len(values)
    if n == 0:
        return 0.0, 0.0, 0.0
    m = _mean(values)
    sd = _stdev(values) if n >= 2 else 0.0
    se = sd / math.sqrt(n) if n >= 2 else 0.0
    return m, sd, se


def build_batch_coverage_workbook(batch):
    """Build the CPCE-style coverage workbook for one batch. Returns a BytesIO."""
    images = list(batch.images.all())
    n = len(images)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data summary'

    first_frame = 2
    last_frame = first_frame + n - 1 if n else first_frame
    mean_col = last_frame + 1
    sd_col = mean_col + 1
    se_col = sd_col + 1

    def cell(r, c, value=None, font=_PLAIN, align=_LEFT, fmt=None, fill=None, border=False):
        cc = ws.cell(row=r, column=c, value=value)
        cc.font = font
        cc.alignment = align
        if fmt:
            cc.number_format = fmt
        if fill:
            cc.fill = PatternFill('solid', fgColor=fill)
        if border:
            cc.border = _BORDER
        return cc

    # ---- Metadata block (left label/value pairs + a right-hand pair) ----
    def right(r, label, value):
        cell(r, mean_col, label, font=_LABEL)
        cell(r, sd_col, value)

    cell(1, 1, 'Project:', font=_LABEL);          cell(1, 2, batch.area_name)
    right(1, 'Analysis date:',
          batch.survey_date.strftime('%Y-%m-%d') if batch.survey_date else '')
    cell(2, 1, 'Dataset name:', font=_LABEL);      cell(2, 2, batch.name)
    right(2, 'Analysis by:', batch.surveyor_names or '')
    cell(3, 1, 'Location:', font=_LABEL);          cell(3, 2, batch.area_name)
    right(3, 'Codefile:', '')
    cell(4, 1, 'Lat / Long:', font=_LABEL)
    cell(4, 2, f'{batch.latitude}, {batch.longitude}')
    cell(5, 1, 'File/sheetname:', font=_LABEL)
    cell(5, 2, f'{_safe_filename(batch)}:Data summary')

    # ---- Transect summary box ----
    r = 7
    box_top = r

    cell(r, 1, 'TRANSECT NAME', font=_HEADER, border=True)
    for i, image in enumerate(images):
        name = image.image.name.split('/')[-1] if image.image else f'Frame {i + 1}'
        cell(r, first_frame + i, name, font=_HEADER, align=_CENTER, border=True)
    r += 1

    cell(r, 1, 'Number of frames', font=_LABEL, border=True)
    for i in range(n):
        cell(r, first_frame + i, 1, align=_CENTER, fmt='0', border=True)
    r += 1

    series, totals = _pct_series(images)

    cell(r, 1, 'Total points', font=_LABEL, border=True)
    for i in range(n):
        cell(r, first_frame + i, totals[i], align=_CENTER, fmt='0', border=True)
    box_bottom = r
    r += 2

    # ---- Major category header row (also carries MEAN / STD. DEV. / STD. ERROR) ----
    hdr = r
    cell(hdr, 1, 'MAJOR CATEGORY (% of transect)', font=_HEADER, fill=HEADER_GRAY, border=True)
    for i in range(n):
        cell(hdr, first_frame + i, None, fill=HEADER_GRAY, border=True)
    cell(hdr, mean_col, 'MEAN', font=_HEADER, align=_CENTER, fill=HEADER_GRAY, border=True)
    cell(hdr, sd_col, 'STD. DEV.', font=_HEADER, align=_CENTER, fill=HEADER_GRAY, border=True)
    cell(hdr, se_col, 'STD. ERROR', font=_HEADER, align=_CENTER, fill=HEADER_GRAY, border=True)
    r += 1

    # ---- One row per category (zebra shaded like the sample) ----
    for idx, name in enumerate(CLASS_ORDER):
        zebra = GRAY if idx % 2 == 0 else None
        cell(r, 1, CATEGORY_DISPLAY[name], font=_PLAIN, fill=zebra, border=True)
        for i in range(n):
            cell(r, first_frame + i, series[name][i], align=_CENTER, fmt=NUM_FMT,
                 fill=zebra, border=True)
        m, sd, se = _stats(series[name])
        cell(r, mean_col, m, align=_CENTER, fmt=NUM_FMT, fill=zebra, border=True)
        cell(r, sd_col, sd, align=_CENTER, fmt=NUM_FMT, fill=zebra, border=True)
        cell(r, se_col, se, align=_CENTER, fmt=NUM_FMT, fill=zebra, border=True)
        r += 1

    # ---- Sum row (the 7 categories cover every point, so each column = 100%) ----
    cell(r, 1, 'Sum (excluding tape+shadow+wand)', font=_LABEL, border=True)
    for i in range(n):
        col_sum = sum(series[name][i] for name in CLASS_ORDER)
        cell(r, first_frame + i, col_sum, font=_LABEL, align=_CENTER, fmt=NUM_FMT, border=True)

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 34
    for c in range(first_frame, last_frame + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    for c in (mean_col, sd_col, se_col):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = ws.cell(row=hdr + 1, column=first_frame)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
